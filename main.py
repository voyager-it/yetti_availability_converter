from tabulate import tabulate
from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# Used to select the workbook in question.
# Make sure to have the sheet for processing set as active and the workbook saved.
Tk().withdraw()
filename = askopenfilename()

wb = load_workbook(filename)
ws = wb.active

# This is for keeping track of the last manually entered row since there isn't a consistent way to track it.
print('Enter final row of selectable range. (2nd row after last category)')
lastRow = int(input())

# header for end table
fullTable = [['Model', 'category', 'dummy_part', 'part', 'description', 'optional']]

i = 2
while i <= lastRow:
    colCount = 5
    colVal = 'temp'
    while colVal is not None:
        colVal = ws.cell(i, colCount).value
        colCount += 1
    colCount -= 2
    j = 6
    while j <= colCount:
        offsetRow = i + 1
        if j == 6:
            optional = 'selected_optional'
        else:
            optional = 'Optional'
        if 'REV A' in str(ws.cell(offsetRow, j).value):
            part = ws.cell(offsetRow, j).value[0:7]
            curRow = [
                ws.cell(1, 1).value,
                ws.cell(i, 2).value,
                ws.cell(i, 1).value,
                part,
                ws.cell(i, j).value,
                optional
            ]
            fullTable.append(curRow)    
        elif 'REV B' not in str(ws.cell(offsetRow, j).value):
            if 'N/A' not in str(ws.cell(offsetRow, j).value):
                if ws.cell(offsetRow, j).value:
                    curRow = [
                        ws.cell(1, 1).value,
                        ws.cell(i, 2).value,
                        ws.cell(i, 1).value,
                        ws.cell(offsetRow, j).value,
                        ws.cell(i, j).value,
                        optional
                    ]
                    fullTable.append(curRow)
        j += 1
    i += 3
# print('done')
# print(tabulate(fullTable))

tempWb = Workbook()
dest_filename = ws.cell(1, 1).value + '.xlsx'

ws1 = tempWb.active
ws1.title = ws.cell(1, 1).value

col = 0
row = 0
for x in fullTable:
    col = 0
    for y in x:
        # ws1[row, col] = fullTable[row][col]
        ws1.cell(column=col+1, row=row+1).value = fullTable[row][col]
        # print(fullTable[row][col])
        col += 1
    row += 1

tempWb.save(filename=dest_filename)
